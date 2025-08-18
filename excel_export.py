# -*- coding: utf-8 -*-
"""
نظام تصدير Excel للتقارير
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO

class ExcelExporter:
    """فئة تصدير البيانات إلى Excel"""
    
    def __init__(self):
        self.currency_symbol = "د.ج"
    
    def export_sales_report(self, sales, start_date=None, end_date=None):
        """تصدير تقرير المبيعات إلى Excel"""
        
        # إنشاء workbook جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير المبيعات"
        
        # إعداد الأنماط
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        title_font = Font(bold=True, size=16)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # العنوان الرئيسي
        ws.merge_cells('A1:H1')
        ws['A1'] = "تقرير المبيعات - متجر الهواتف المحمولة"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # فترة التقرير
        if start_date and end_date:
            ws.merge_cells('A2:H2')
            ws['A2'] = f"من {start_date.strftime('%Y-%m-%d')} إلى {end_date.strftime('%Y-%m-%d')}"
            ws['A2'].alignment = title_alignment
            header_row = 4
        else:
            header_row = 3
        
        # رؤوس الأعمدة
        headers = [
            'رقم الفاتورة', 'التاريخ', 'العميل', 'المجموع الفرعي', 
            'الخصم', 'المجموع النهائي', 'طريقة الدفع', 'ملاحظات'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # بيانات المبيعات
        total_sales = 0
        total_discount = 0
        total_final = 0
        
        for row, sale in enumerate(sales, header_row + 1):
            ws.cell(row=row, column=1, value=sale.id).border = border
            ws.cell(row=row, column=2, value=sale.created_at.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=3, value=sale.customer.name if sale.customer else 'عميل غير محدد').border = border
            ws.cell(row=row, column=4, value=f"{sale.total_amount:.2f} {self.currency_symbol}").border = border
            ws.cell(row=row, column=5, value=f"{sale.discount:.2f} {self.currency_symbol}").border = border
            ws.cell(row=row, column=6, value=f"{sale.final_amount:.2f} {self.currency_symbol}").border = border
            ws.cell(row=row, column=7, value=sale.payment_method).border = border
            ws.cell(row=row, column=8, value=sale.notes or '').border = border
            
            total_sales += sale.total_amount
            total_discount += sale.discount
            total_final += sale.final_amount
        
        # إجمالي المبيعات
        summary_row = len(sales) + header_row + 2
        ws.merge_cells(f'A{summary_row}:C{summary_row}')
        ws[f'A{summary_row}'] = "الإجمالي:"
        ws[f'A{summary_row}'].font = Font(bold=True)
        ws[f'A{summary_row}'].alignment = Alignment(horizontal="center")
        
        ws[f'D{summary_row}'] = f"{total_sales:.2f} {self.currency_symbol}"
        ws[f'E{summary_row}'] = f"{total_discount:.2f} {self.currency_symbol}"
        ws[f'F{summary_row}'] = f"{total_final:.2f} {self.currency_symbol}"
        
        for col in range(1, 9):
            ws.cell(row=summary_row, column=col).border = border
            ws.cell(row=summary_row, column=col).font = Font(bold=True)
        
        # تنسيق عرض الأعمدة
        column_widths = [12, 12, 20, 15, 12, 15, 12, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # حفظ في buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_products_report(self, products):
        """تصدير تقرير المنتجات إلى Excel"""
        
        # إنشاء workbook جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير المنتجات"
        
        # إعداد الأنماط
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        title_font = Font(bold=True, size=16)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # العنوان الرئيسي
        ws.merge_cells('A1:J1')
        ws['A1'] = "تقرير المنتجات - متجر الهواتف المحمولة"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # رؤوس الأعمدة
        headers = [
            'الرقم', 'اسم المنتج', 'الماركة', 'الموديل', 'اللون',
            'سعر الشراء', 'سعر البيع', 'الكمية', 'الحد الأدنى', 'الفئة'
        ]
        
        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # بيانات المنتجات
        total_value = 0
        low_stock_count = 0
        
        for row, product in enumerate(products, header_row + 1):
            ws.cell(row=row, column=1, value=product.id).border = border
            ws.cell(row=row, column=2, value=product.name).border = border
            ws.cell(row=row, column=3, value=product.brand).border = border
            ws.cell(row=row, column=4, value=product.model).border = border
            ws.cell(row=row, column=5, value=product.color or '').border = border
            ws.cell(row=row, column=6, value=f"{product.price_buy:.2f} {self.currency_symbol}").border = border
            ws.cell(row=row, column=7, value=f"{product.price_sell:.2f} {self.currency_symbol}").border = border
            
            # تلوين الكمية إذا كانت منخفضة
            quantity_cell = ws.cell(row=row, column=8, value=product.quantity)
            quantity_cell.border = border
            if product.quantity <= product.min_quantity:
                quantity_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                low_stock_count += 1
            
            ws.cell(row=row, column=9, value=product.min_quantity).border = border
            ws.cell(row=row, column=10, value=product.category.name if product.category else '').border = border
            
            total_value += product.price_buy * product.quantity
        
        # إحصائيات
        summary_row = len(products) + header_row + 2
        ws.merge_cells(f'A{summary_row}:D{summary_row}')
        ws[f'A{summary_row}'] = "إحصائيات المخزون:"
        ws[f'A{summary_row}'].font = Font(bold=True, size=12)
        
        ws[f'A{summary_row + 1}'] = f"إجمالي المنتجات: {len(products)}"
        ws[f'A{summary_row + 2}'] = f"منتجات منخفضة المخزون: {low_stock_count}"
        ws[f'A{summary_row + 3}'] = f"قيمة المخزون الإجمالية: {total_value:.2f} {self.currency_symbol}"
        
        # تنسيق عرض الأعمدة
        column_widths = [8, 25, 15, 15, 12, 15, 15, 10, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # حفظ في buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_purchase_report(self, purchases, start_date=None, end_date=None):
        """تصدير تقرير المشتريات إلى Excel"""
        
        # إنشاء workbook جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير المشتريات"
        
        # إعداد الأنماط
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        title_font = Font(bold=True, size=16)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # العنوان الرئيسي
        ws.merge_cells('A1:H1')
        ws['A1'] = "تقرير المشتريات - متجر الهواتف المحمولة"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # فترة التقرير
        if start_date and end_date:
            ws.merge_cells('A2:H2')
            ws['A2'] = f"من {start_date.strftime('%Y-%m-%d')} إلى {end_date.strftime('%Y-%m-%d')}"
            ws['A2'].alignment = title_alignment
            header_row = 4
        else:
            header_row = 3
        
        # رؤوس الأعمدة
        headers = [
            'رقم الفاتورة', 'التاريخ', 'المورد', 'المجموع الفرعي', 
            'الخصم', 'المجموع النهائي', 'طريقة الدفع', 'ملاحظات'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # بيانات المشتريات
        total_purchases = 0
        total_discount = 0
        total_final = 0
        
        for row, purchase in enumerate(purchases, header_row + 1):
            ws.cell(row=row, column=1, value=purchase.invoice_number).border = border
            ws.cell(row=row, column=2, value=purchase.created_at.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=3, value=purchase.supplier.name).border = border
            ws.cell(row=row, column=4, value=f"{purchase.total_amount:.2f} {self.currency_symbol}").border = border
            ws.cell(row=row, column=5, value=f"{purchase.discount:.2f} {self.currency_symbol}").border = border
            ws.cell(row=row, column=6, value=f"{purchase.final_amount:.2f} {self.currency_symbol}").border = border
            ws.cell(row=row, column=7, value=purchase.payment_method).border = border
            ws.cell(row=row, column=8, value=purchase.notes or '').border = border
            
            total_purchases += purchase.total_amount
            total_discount += purchase.discount
            total_final += purchase.final_amount
        
        # إجمالي المشتريات
        summary_row = len(purchases) + header_row + 2
        ws.merge_cells(f'A{summary_row}:C{summary_row}')
        ws[f'A{summary_row}'] = "الإجمالي:"
        ws[f'A{summary_row}'].font = Font(bold=True)
        ws[f'A{summary_row}'].alignment = Alignment(horizontal="center")
        
        ws[f'D{summary_row}'] = f"{total_purchases:.2f} {self.currency_symbol}"
        ws[f'E{summary_row}'] = f"{total_discount:.2f} {self.currency_symbol}"
        ws[f'F{summary_row}'] = f"{total_final:.2f} {self.currency_symbol}"
        
        for col in range(1, 9):
            ws.cell(row=summary_row, column=col).border = border
            ws.cell(row=summary_row, column=col).font = Font(bold=True)
        
        # تنسيق عرض الأعمدة
        column_widths = [15, 12, 20, 15, 12, 15, 12, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # حفظ في buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

def format_currency_excel(amount):
    """تنسيق العملة للـ Excel"""
    return f"{amount:,.2f} د.ج"